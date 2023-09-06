import json
from collections import defaultdict
from dataclasses import dataclass, astuple as _astuple
from pathlib import Path
from statistics import mean, stdev
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import get_column_letter
import numpy as np
from copy import copy


def astuple(self):
    return _astuple(self)


games_dir = Path(__file__).parent.parent / "processed"
total_games = len(list(games_dir.glob("*.json")))

data = []
for game_file in games_dir.glob("*.json"):
    with game_file.open() as f:
        data.extend(json.load(f)["rows"])

hero_tiers = {
    "Anti-Mage": 1,
    "Batrider": 1,
    "Bounty Hunter": 1,
    "Crystal Maiden": 1,
    "Dazzle": 1,
    "Drow Ranger": 1,
    "Enchantress": 1,
    "Lich": 1,
    "Magnus": 1,
    "Phantom Assassin": 1,
    "Shadow Demon": 1,
    "Slardar": 1,
    "Snapfire": 1,
    "Tusk": 1,
    "Vengeful Spirit": 1,
    "Venomancer": 1,
    # T2
    "Bristleback": 2,
    "Chaos Knight": 2,
    "Earth Spirit": 2,
    "Juggernaut": 2,
    "Kunkka": 2,
    "Legion Commander": 2,
    "Luna": 2,
    "Meepo": 2,
    "Nature's Prophet": 2,
    "Pudge": 2,
    "Queen of Pain": 2,
    "Spirit Breaker": 2,
    "Storm Spirit": 2,
    "Windranger": 2,
    # T3
    "Abaddon": 3,
    "Alchemist": 3,
    "Beastmaster": 3,
    "Ember Spirit": 3,
    "Lifestealer": 3,
    "Lycan": 3,
    "Omniknight": 3,
    "Puck": 3,
    "Shadow Shaman": 3,
    "Slark": 3,
    "Spectre": 3,
    "Terrorblade": 3,
    "Treant Protector": 3,
    # T4
    "Death Prophet": 4,
    "Doom": 4,
    "Lina": 4,
    "Lone Druid": 4,
    "Mirana": 4,
    "Pangolier": 4,
    "Rubick": 4,
    "Sven": 4,
    "Templar Assassin": 4,
    "Tidehunter": 4,
    "Viper": 4,
    "Void Spirit": 4,
    # T5
    "Axe": 5,
    "Dragon Knight": 5,
    "Faceless Void": 5,
    "Keeper of the Light": 5,
    "Medusa": 5,
    "Troll Warlord": 5,
    "Wraith King": 5,
}

item_tiers = {
    "Crown of Antlers": 3,
    "Armlet of Mordiggian": 2,
    "Skull Basher": 3,
    "Battle Fury": 3,
    "Black King Bar": 4,
    "Blade Mail": 2,
    "Bloodthorn": 5,
    "Arcane Boots": 2,
    "Butterfly": 4,
    "Chainmail": 1,
    "Claymore": 1,
    "Craggy Coat": 3,
    "Blink Dagger": 2,
    "Dagon": 4,
    "Desolator": 2,
    "Diffusal Blade": 4,
    "Eul's Scepter": 3,
    "Leg Breaker's Fedora": 3,
    "Gloves of Haste": 1,
    "Heaven's Halberd": 2,
    "Headdress": 1,
    "Hood of Defiance": 1,
    "Horn of the Alpha": 5,
    "Kaden's Blade": 4,
    "Kaya": 1,
    "Dragon Lance": 2,
    "Maelstrom": 4,
    "Mekansm": 4,
    "Monkey King Bar": 4,
    "Mask of Madness": 3,
    "Moon Shard": 4,
    "Morbid Mask": 1,
    "Necronomicon": 3,
    "Octarine Essence": 3,
    "Refresher Orb": 4,
    "Paladin Sword": 2,
    "Stonehall Pike": 2,
    "Pipe of Insight": 3,
    "Pirate Hat": 3,
    "Quelling Blade": 2,
    "Radiance": 5,
    "Divine Rapier": 5,
    "Ristul Circlet": 3,
    "Satanic": 5,
    "Scythe of Vyse": 4,
    "Witless Shako": 3,
    "Shiva's Guard": 5,
    "Silver Edge": 3,
    "Eye of Skadi": 4,
    "Stonehall Cloak": 2,
    "Talisman of Evasion": 1,
    "Heart of Tarrasque": 5,
    "Vanguard": 2,
    "Vitality Booster": 1,
    "Vesture of the Tyrant": 5,
    "Vladmir's Offering": 5,
    "Void Stone": 1,
}

tier_heroes_seen_by_stars = {
    1: {1: 0, 2: 0, 3: 0, "total": 0},
    2: {1: 0, 2: 0, 3: 0, "total": 0},
    3: {1: 0, 2: 0, 3: 0, "total": 0},
    4: {1: 0, 2: 0, 3: 0, "total": 0},
    5: {1: 0, 2: 0, 3: 0, "total": 0},
}

hero_finishes = defaultdict(lambda: {1: [], 2: [], 3: []})
underlord_finishes = defaultdict(list)
item_finishes = defaultdict(list)
alliance_finishes = defaultdict(list)
active_alliance_finishes = defaultdict(list)
alliance_pair_finishes = defaultdict(lambda: defaultdict(list))
hero_finishes_by_rank = defaultdict(lambda: defaultdict(int))
hero_star_finishes_by_rank = defaultdict(lambda: defaultdict(list))
underlord_finishes_by_rank = defaultdict(lambda: defaultdict(int))
alliance_finishes_by_rank = defaultdict(lambda: defaultdict(int))
active_alliance_finishes_by_rank = defaultdict(lambda: defaultdict(int))
item_finishes_by_rank = defaultdict(lambda: defaultdict(int))
finishes_by_rank = defaultdict(int)


def active_alliances(alliances):
    yield from (
        (name, alliance)
        for name, alliance in alliances.items()
        if alliance["active"]
    )


for row in data:
    if not row["underlord"]:
        continue
    underlord_finishes[row["underlord"]].append(row["rank"])
    underlord_finishes[row["underlord"].split(" ")[0]].append(row["rank"])
    underlord_finishes_by_rank[row["rank"]][row["underlord"]] += 1
    underlord_finishes_by_rank[row["rank"]][
        row["underlord"].split(" ")[0]
    ] += 1
    finishes_by_rank[row["rank"]] += 1
    for name, alliance_data in active_alliances(row["alliances"]):
        alliance_finishes[name].append(row["rank"])
        active_alliance_finishes[(name, alliance_data["active"])].append(
            row["rank"]
        )
        alliance_finishes_by_rank[row["rank"]][name] += 1
        active_alliance_finishes_by_rank[row["rank"]][
            (name, alliance_data["active"])
        ] += 1
        for name2, data2 in active_alliances(row["alliances"]):
            # Intentionally don't skip if name == name2
            alliance_pair_finishes[name][name2].append(row["rank"])
    for hero in row["heroes"]:
        tier_heroes_seen_by_stars[hero_tiers[hero["name"]]][hero["stars"]] += 1
        tier_heroes_seen_by_stars[hero_tiers[hero["name"]]]["total"] += 1
        hero_finishes[hero["name"]][hero["stars"]].append(row["rank"])
        hero_finishes_by_rank[row["rank"]][hero["name"]] += 1
        hero_star_finishes_by_rank[row["rank"]][hero["name"]].append(
            hero["stars"]
        )
        if hero["item"]:
            item_finishes[hero["item"]].append(row["rank"])
            item_finishes_by_rank[row["rank"]][hero["item"]] += 1

finish_rates_by_tier = {
    n: {
        1: tier_heroes_seen_by_stars[n][1]
        / tier_heroes_seen_by_stars[n]["total"],
        2: tier_heroes_seen_by_stars[n][2]
        / tier_heroes_seen_by_stars[n]["total"],
        3: tier_heroes_seen_by_stars[n][3]
        / tier_heroes_seen_by_stars[n]["total"],
        "total": tier_heroes_seen_by_stars[n]["total"] / len(data),
    }
    for n in range(1, 6)
}
hero_finish_f_by_rank = {
    rank: {
        hero: hero_fs / finishes_by_rank[rank]
        for hero, hero_fs in finishes.items()
    }
    for rank, finishes in hero_finishes_by_rank.items()
}


@dataclass
class HeroRow:
    name: str
    total_pickrate: float
    pickrate_1: float
    p_1: float
    pickrate_2: float
    p_2: float
    pickrate_3: float
    p_3: float

    avg_finish_1: float
    finish_std_dev_1: float
    avg_finish_2: float
    finish_std_dev_2: float
    avg_finish_3: float
    finish_std_dev_3: float
    avg_finish_total: float
    finish_std_dev_total: float

    rank_slope: float
    tier_normalized_rank_slope: float

    @property
    def tier(self):
        return hero_tiers[self.name]

    @property
    def weighted_avg_finish(self):
        return (
            (self.avg_finish_1 or 0)
            * (finish_rates_by_tier[self.tier][1] or 0)
            + (self.avg_finish_2 or 0)
            * (finish_rates_by_tier[self.tier][2] or 0)
            + (self.avg_finish_3 or 0)
            * (finish_rates_by_tier[self.tier][3] or 0)
        )

    @staticmethod
    def header():
        return [
            "Hero",
            "Tier",
            "Pick Rate (f)",
            "f[⭐]",
            "P(⭐)",
            "f[⭐⭐]",
            "P(⭐⭐)",
            "f[⭐⭐⭐]",
            "P(⭐⭐⭐)",
            "⭐ Mean Rank",
            "Std Dev",
            "⭐⭐ Mean Rank",
            "Std Dev",
            "⭐⭐⭐ Mean Rank",
            "Std Dev",
            "Mean Rank",
            "Std Dev",
            "Tier-normalized Mean Rank",
            "Rank Slope",
            "Tier-normalized Rank Slope"
        ]

    def to_row(self):
        return [
            self.name,
            self.tier,
            self.total_pickrate,
            self.pickrate_1,
            self.p_1,
            self.pickrate_2,
            self.p_2,
            self.pickrate_3,
            self.p_3,
            self.avg_finish_1,
            self.finish_std_dev_1,
            self.avg_finish_2,
            self.finish_std_dev_2,
            self.avg_finish_3,
            self.finish_std_dev_3,
            self.avg_finish_total,
            self.finish_std_dev_total,
            self.weighted_avg_finish,
            self.rank_slope,
            self.tier_normalized_rank_slope,
        ]


@dataclass
class UnderlordRow:
    name: str
    pickrate: float
    avg_finish: float
    finish_std_dev: float
    rank_slope: float

    to_row = astuple

    @property
    def is_total(self):
        return self.name.split(" ")[0] == self.name

    @staticmethod
    def header():
        return [
            "Underlord",
            "Pick Rate",
            "Mean Rank",
            "Std Dev",
            "Rank Slope",
        ]


@dataclass
class AllianceRow:
    name: str
    pickrate: float
    avg_finish: float
    finish_std_dev: float
    rank_slope: float

    to_row = astuple

    @staticmethod
    def header():
        return [
            "Alliance",
            "Pick Rate",
            "Mean Rank",
            "Std Dev",
            "Rank Slope",
        ]


@dataclass
class ActiveAllianceRow:
    name: str
    level: int
    pickrate: float
    avg_finish: float
    finish_std_dev: float
    rank_slope: float

    to_row = astuple

    @staticmethod
    def header():
        return [
            "Alliance",
            "Level",
            "Pick Rate",
            "Mean Rank",
            "Std Dev",
            "Rank Slope",
        ]


@dataclass
class ItemRow:
    name: str
    pickrate: float
    avg_finish: float
    finish_std_dev: float
    rank_slope: float

    @property
    def tier(self):
        return item_tiers[self.name]

    @staticmethod
    def header():
        return [
            "Item",
            "Tier",
            "Pick Rate",
            "Mean Rank",
            "Std Dev",
            "Rank Slope",
        ]

    def to_row(self):
        return [
            self.name,
            self.tier,
            self.pickrate,
            self.avg_finish,
            self.finish_std_dev,
            self.rank_slope,
        ]


higher_is_better_rule = ColorScaleRule(
    start_type="min",
    start_color="FFAAAA",
    mid_type="percentile",
    mid_value=50,
    mid_color="FFFFFF",
    end_type="max",
    end_color="AAAAFF",
)
lower_is_better_rule = ColorScaleRule(
    start_type="min",
    start_color="AAAAFF",
    mid_type="percentile",
    mid_value=50,
    mid_color="FFFFFF",
    end_type="max",
    end_color="FFAAAA",
)
slope_rule = ColorScaleRule(
    start_type="min",
    start_color="AAAAFF",
    mid_value=0,
    mid_type="num",
    mid_color="FFFFFF",
    end_type="max",
    end_color="FFAAAA",
)


@dataclass
class UnderlordsMeta:
    games_analyzed: int
    rows_analyzed: int
    tier_star_frequencies: dict[int, dict[int, float]]
    hero_stats: list[HeroRow]
    underlord_stats: list[UnderlordRow]
    alliance_stats: list[AllianceRow]
    active_alliance_stats: list[ActiveAllianceRow]
    alliance_pair_scores: dict[str, dict[str, float]]
    item_stats: list[ItemRow]

    def to_workbook(self):
        workbook = Workbook()
        ws: Worksheet = workbook.active

        # Meta
        ws.title = "Meta"
        ws.append(["Games Analyzed", self.games_analyzed])
        ws.append(["Rows Analyzed", self.rows_analyzed])

        # Hero scores
        ws = workbook.create_sheet("Hero Scores")
        ws.append(HeroRow.header())
        for row in sorted(
            self.hero_stats,
            key=lambda h: (h.tier, h.tier_normalized_rank_slope),
        ):
            ws.append(row.to_row())
        for col in ["C", "D", "E", "F", "G", "H", "I"]:
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                higher_is_better_rule,
            )

        for col in ["L", "J", "K", "M", "N", "O", "P", "Q", "R"]:
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                lower_is_better_rule,
            )
        for col in "ST":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                slope_rule,
            )
        for col in ws.iter_cols(min_row=2, min_col=10, max_col=18):
            for cell in col:
                cell.number_format = "0.00"
        for col in ws.iter_cols(min_row=2, min_col=3, max_col=9):
            for cell in col:
                cell.number_format = "0.00%"
        for col in ws.iter_cols(min_row=2, min_col=19):
            for cell in col:
                cell.number_format = "0.00%"
        ws.freeze_panes = "C2"

        # Underlord scores
        ws = workbook.create_sheet("Underlord Scores")
        ws.append(UnderlordRow.header())
        for row in sorted(self.underlord_stats, key=lambda u: u.rank_slope):
            ws.append(row.to_row())
            if row.is_total:
                bold_font = copy(ws["A"][-1].font)
                bold_font.bold = True
                ws["A"][-1].font = bold_font
        for col in "B":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                higher_is_better_rule,
            )
        for col in "CD":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                lower_is_better_rule,
            )
        for col in "E":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                slope_rule,
            )
        for row in ws.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.number_format = "0.00"
        for col in "BE":
            for cell in ws[col][1:]:
                cell.number_format = "0.00%"
        ws.freeze_panes = "B2"

        # Alliance scores
        ws = workbook.create_sheet("Alliance Scores")
        ws.append(AllianceRow.header())
        for row in sorted(self.alliance_stats, key=lambda a: a.rank_slope):
            ws.append(row.to_row())
        for col in "B":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                higher_is_better_rule,
            )
        for col in "CD":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                lower_is_better_rule,
            )
        for col in "E":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                slope_rule,
            )
        for row in ws.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.number_format = "0.00"
        for col in "BE":
            for cell in ws[col][1:]:
                cell.number_format = "0.00%"
        ws.freeze_panes = "B2"

        # Active Alliance scores
        ws = workbook.create_sheet("Active Alliance Scores")
        ws.append(ActiveAllianceRow.header())
        for row in sorted(
            self.active_alliance_stats, key=lambda a: a.rank_slope
        ):
            ws.append(row.to_row())
        for col in "C":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                higher_is_better_rule,
            )
        for col in "DE":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                lower_is_better_rule,
            )
        for col in "F":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                slope_rule,
            )
        for row in ws.iter_rows(min_row=2, min_col=4):
            for cell in row:
                cell.number_format = "0.00"
        for col in "CF":
            for cell in ws[col][1:]:
                cell.number_format = "0.00%"
        # TODO normalize by tier
        ws.freeze_panes = "C2"

        # Item scores
        ws = workbook.create_sheet("Item Scores")
        ws.append(ItemRow.header())
        for row in sorted(
            self.item_stats, key=lambda i: (i.tier, i.rank_slope)
        ):
            ws.append(row.to_row())
        for col in "C":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                higher_is_better_rule,
            )
        for col in "DE":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                lower_is_better_rule,
            )
        for col in "F":
            ws.conditional_formatting.add(
                f"{col}2:{col}{len(self.hero_stats) + 1}",
                slope_rule,
            )
        for col in ws.iter_cols(min_row=2, min_col=4):
            for cell in col:
                cell.number_format = "0.00"
        for col in "CF":
            for cell in ws[col][1:]:
                cell.number_format = "0.00%"
        ws.freeze_panes = "C2"

        # Alliance Pair scores
        ws = workbook.create_sheet("Alliance Pair Scores")
        ws.append([""] + sorted(self.alliance_pair_scores))
        for name in sorted(self.alliance_pair_scores):
            ws.append(
                [name]
                + [
                    v
                    for _, v in sorted(self.alliance_pair_scores[name].items())
                ]
            )
        ws.conditional_formatting.add(
            f"B2:{get_column_letter(ws.max_column)}{ws.max_row}",
            lower_is_better_rule,
        )
        for col in ws.iter_cols(min_row=2, min_col=2):
            for cell in col:
                cell.number_format = "0.00"
        ws.freeze_panes = "B2"

        return workbook


def build_stats():
    tier_star_frequencies = finish_rates_by_tier
    hero_stats = [
        HeroRow(
            name=hero,
            total_pickrate=(
                len(hero_finishes[hero][1])
                + len(hero_finishes[hero][2])
                + len(hero_finishes[hero][3])
            )
            / len(data),
            pickrate_1=len(hero_finishes[hero][1]) / len(data),
            p_1=len(hero_finishes[hero][1])
            / (
                len(hero_finishes[hero][1])
                + len(hero_finishes[hero][2])
                + len(hero_finishes[hero][3])
            ),
            pickrate_2=len(hero_finishes[hero][2]) / len(data),
            p_2=len(hero_finishes[hero][2])
            / (
                len(hero_finishes[hero][1])
                + len(hero_finishes[hero][2])
                + len(hero_finishes[hero][3])
            ),
            pickrate_3=len(hero_finishes[hero][3]) / len(data),
            p_3=len(hero_finishes[hero][3])
            / (
                len(hero_finishes[hero][1])
                + len(hero_finishes[hero][2])
                + len(hero_finishes[hero][3])
            ),
            avg_finish_1=mean(hero_finishes[hero][1])
            if hero_finishes[hero][1]
            else None,
            finish_std_dev_1=stdev(hero_finishes[hero][1])
            if len(hero_finishes[hero][1]) > 1
            else None,
            avg_finish_2=mean(hero_finishes[hero][2])
            if hero_finishes[hero][2]
            else None,
            finish_std_dev_2=stdev(hero_finishes[hero][2])
            if len(hero_finishes[hero][2]) > 1
            else None,
            avg_finish_3=mean(hero_finishes[hero][3])
            if hero_finishes[hero][3]
            else None,
            finish_std_dev_3=stdev(hero_finishes[hero][3])
            if len(hero_finishes[hero][3]) > 1
            else None,
            avg_finish_total=mean(
                hero_finishes[hero][1]
                + hero_finishes[hero][2]
                + hero_finishes[hero][3]
            ),
            finish_std_dev_total=stdev(
                hero_finishes[hero][1]
                + hero_finishes[hero][2]
                + hero_finishes[hero][3]
            ),
            rank_slope=float(
                np.polyfit(
                    *zip(
                        *[
                            (rank, hero_finish_f_by_rank[rank][hero])
                            for rank in range(1, 9)
                            if hero_finish_f_by_rank[rank].get(hero, None)
                        ]
                    ),
                    1,
                )[0]
            ),
            tier_normalized_rank_slope=float(
                np.polyfit(
                    *zip(
                        *[
                            (
                                rank,
                                sum(
                                    finish_rates_by_tier[hero_tiers[hero]][
                                        star
                                    ]
                                    for star in hero_star_finishes_by_rank[
                                        rank
                                    ][hero]
                                )
                                / finishes_by_rank[rank],
                            )
                            for rank in range(1, 9)
                            if hero_star_finishes_by_rank[rank].get(hero, None)
                        ]
                    ),
                    1,
                )[0]
            ),
        )
        for hero in hero_finishes
    ]
    underlord_stats = [
        UnderlordRow(
            name=underlord,
            pickrate=len(underlord_finishes[underlord]) / len(data),
            avg_finish=mean(underlord_finishes[underlord]),
            finish_std_dev=stdev(underlord_finishes[underlord])
            if len(underlord_finishes[underlord]) > 1
            else None,
            rank_slope=float(
                np.polyfit(
                    *zip(
                        *[
                            (
                                rank,
                                underlord_finishes_by_rank[rank][underlord]
                                / finishes_by_rank[rank],
                            )
                            for rank in range(1, 9)
                            if underlord_finishes_by_rank[rank].get(
                                underlord, None
                            )
                        ]
                    ),
                    1,
                )[0]
            ),
        )
        for underlord in underlord_finishes
    ]
    alliance_stats = [
        AllianceRow(
            name=alliance,
            pickrate=len(alliance_finishes[alliance]) / len(data),
            avg_finish=mean(alliance_finishes[alliance]),
            finish_std_dev=stdev(alliance_finishes[alliance])
            if len(alliance_finishes[alliance]) > 1
            else None,
            rank_slope=float(
                np.polyfit(
                    *zip(
                        *[
                            (
                                rank,
                                alliance_finishes_by_rank[rank][alliance]
                                / finishes_by_rank[rank],
                            )
                            for rank in range(1, 9)
                            if alliance_finishes_by_rank[rank].get(
                                alliance, None
                            )
                        ]
                    ),
                    1,
                )[0]
            ),
        )
        for alliance in alliance_finishes
    ]
    active_alliance_stats = [
        ActiveAllianceRow(
            name=alliance[0],
            level=alliance[1],
            pickrate=len(active_alliance_finishes[alliance]) / len(data),
            avg_finish=mean(active_alliance_finishes[alliance]),
            finish_std_dev=stdev(active_alliance_finishes[alliance])
            if len(active_alliance_finishes[alliance]) > 1
            else None,
            rank_slope=float(
                np.polyfit(
                    *zip(
                        *[
                            (
                                rank,
                                active_alliance_finishes_by_rank[rank][
                                    alliance
                                ]
                                / finishes_by_rank[rank],
                            )
                            for rank in range(1, 9)
                            if active_alliance_finishes_by_rank[rank].get(
                                alliance, None
                            )
                        ]
                    ),
                    1,
                )[0]
            ),
        )
        for alliance in active_alliance_finishes
    ]
    item_stats = [
        ItemRow(
            name=item,
            pickrate=len(item_finishes[item]) / len(data),
            avg_finish=mean(item_finishes[item]),
            finish_std_dev=stdev(item_finishes[item])
            if len(item_finishes[item]) > 1
            else None,
            rank_slope=float(
                np.polyfit(
                    *zip(
                        *[
                            (
                                rank,
                                item_finishes_by_rank[rank][item]
                                / finishes_by_rank[rank],
                            )
                            for rank in range(1, 9)
                            if item_finishes_by_rank[rank].get(item, None)
                        ]
                    ),
                    1,
                )[0]
            ),
        )
        for item in item_finishes
    ]
    alliance_pair_scores = {
        name: {
            name2: mean(alliance_pair_finishes[name][name2])
            for name2 in alliance_pair_finishes[name]
        }
        for name in alliance_pair_finishes
    }
    return UnderlordsMeta(
        games_analyzed=total_games,
        rows_analyzed=len(data),
        tier_star_frequencies=tier_star_frequencies,
        hero_stats=hero_stats,
        underlord_stats=underlord_stats,
        alliance_stats=alliance_stats,
        active_alliance_stats=active_alliance_stats,
        item_stats=item_stats,
        alliance_pair_scores=alliance_pair_scores,
    )


stats = build_stats()
workbook = stats.to_workbook()
workbook.save("stats.xlsx")
