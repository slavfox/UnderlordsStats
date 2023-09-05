import json
from pathlib import Path
# import average
from statistics import mean, stdev

from openpyxl import Workbook

games_dir = Path(__file__).parent.parent / "processed"
results_dir = Path(__file__).parent / "results"

total_games = len(list(games_dir.glob("*.json")))
hero_finishes = {}
hero_finishes_by_stars = {}
underlord_finishes = {}
alliance_finishes = {}
active_alliance_finishes = {}
item_finishes = {}
hero_finishes_by_item = {}

for game_file in games_dir.glob("*.json"):
    with game_file.open() as f:
        game = json.load(f)
        for row in game["rows"]:
            rank = row["rank"]
            underlord_finishes.setdefault(row["underlord"], []).append(rank)
            for name, alliance in row["alliances"].items():
                if alliance["active"]:
                    alliance_finishes.setdefault(name, []).append(rank)
                    active_alliance_finishes.setdefault(
                        (name, alliance["active"]), []
                    ).append(rank)

            for hero in row["heroes"]:
                hero_finishes.setdefault(hero["name"], []).append(rank)
                hero_finishes_by_stars.setdefault(
                    hero["name"], {1: [], 2: [], 3: []}
                )[hero["stars"]].append(rank)
                hero_finishes_by_item.setdefault(
                    hero["name"],
                    {},
                ).setdefault(
                    hero["item"], []
                ).append(rank)
                if hero["item"]:
                    item_finishes.setdefault(hero["item"], []).append(rank)


workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Hero Avg Finish"
worksheet.append(
    [
        "Hero",
        "Total Games",
        "Pickrate",
        "1 star finishes",
        "1 star pickrate",
        "2 star finishes",
        "2 star pickrate",
        "3 star finishes",
        "3 star pickrate",
        "1 star avg finish",
        "1 star finish std dev",
        "2 star avg finish",
        "2 star finish std dev",
        "3 star avg finish",
        "3 star finish std dev",
        "Avg Finish",
        "Finish std dev",
    ]
)
rows = [
    {
        "name": hero,
        "total": len(hero_finishes[hero]),
        "pickrate": len(hero_finishes[hero]) / total_games,
        "1 star": len(hero_finishes_by_stars[hero][1]),
        "1 star pickrate": len(hero_finishes_by_stars[hero][1]) / total_games,
        "2 star": len(hero_finishes_by_stars[hero][2]),
        "2 star pickrate": len(hero_finishes_by_stars[hero][2]) / total_games,
        "3 star": len(hero_finishes_by_stars[hero][3]),
        "3 star pickrate": len(hero_finishes_by_stars[hero][3]) / total_games,
        "1 star avg finish": sum(hero_finishes_by_stars[hero][1])
        / len(hero_finishes_by_stars[hero][1] or [None]),
        "1 star finish std dev": stdev(hero_finishes_by_stars[hero][1])
        if len(hero_finishes_by_stars[hero][1]) > 1
        else None,
        "2 star avg finish": sum(hero_finishes_by_stars[hero][2])
        / len(hero_finishes_by_stars[hero][2] or [None]),
        "2 star finish std dev": stdev(hero_finishes_by_stars[hero][2])
        if len(hero_finishes_by_stars[hero][2]) > 1
        else None,
        "3 star avg finish": (
            sum(hero_finishes_by_stars[hero][3])
            / len(hero_finishes_by_stars[hero][3] or [None])
        ),
        "3 star finish std dev": stdev(hero_finishes_by_stars[hero][3])
        if len(hero_finishes_by_stars[hero][3]) > 1
        else None,
        "avg finish": sum(hero_finishes[hero])
        / len(hero_finishes[hero] or [None]),
        "finish std dev": stdev(hero_finishes[hero])
        if len(hero_finishes[hero]) > 1
        else None,
    }
    for hero in hero_finishes
]

for row in sorted(rows, key=lambda x: x["avg finish"]):
    worksheet.append(
        [
            row["name"],
            row["total"],
            row["pickrate"],
            row["1 star"],
            row["1 star pickrate"],
            row["2 star"],
            row["2 star pickrate"],
            row["3 star"],
            row["3 star pickrate"],
            row["1 star avg finish"],
            row["1 star finish std dev"],
            row["2 star avg finish"],
            row["2 star finish std dev"],
            row["3 star avg finish"],
            row["3 star finish std dev"],
            row["avg finish"],
            row["finish std dev"],
        ]
    )


worksheet = workbook.create_sheet("Underlord Avg Finish")
worksheet.append(
    ["Underlord", "Total Games", "Avg Finish", "Finish std dev", "Pickrate"]
)
rows = []
for ul in underlord_finishes:
    rows.append(
        {
            "name": ul,
            "total": len(underlord_finishes[ul]),
            "avg": sum(underlord_finishes[ul]) / len(underlord_finishes[ul]),
            "std dev": stdev(underlord_finishes[ul])
            if len(underlord_finishes[ul]) > 1
            else None,
            "pickrate": len(underlord_finishes[ul]) / total_games,
        }
    )

for ul in {ul.split("-")[0] for ul in underlord_finishes if ul}:
    finishes = []
    for u in underlord_finishes:
        if u and u.startswith(ul):
            finishes += underlord_finishes[u]
    rows.append(
        {
            "name": ul,
            "total": len(finishes),
            "avg": sum(finishes) / len(finishes),
            "std dev": stdev(finishes) if len(finishes) > 1 else None,
            "pickrate": len(finishes) / total_games,
        }
    )

for row in sorted(rows, key=lambda x: x["avg"]):
    worksheet.append(
        [
            row["name"],
            row["total"],
            row["avg"],
            row["std dev"],
            row["pickrate"],
        ]
    )

worksheet = workbook.create_sheet("Alliance Avg Finish")
worksheet.append(
    ["Alliance", "Total Games", "Avg Finish", "Std Dev", "Pickrate"]
)
rows = [
    [
        name,
        len(alliance_finishes[name]),
        sum(alliance_finishes[name]) / len(alliance_finishes[name]),
        stdev(alliance_finishes[name])
        if len(alliance_finishes[name]) > 1
        else None,
        len(alliance_finishes[name]) / total_games,
    ]
    for name in alliance_finishes
]

for row in sorted(rows, key=lambda x: x[1]):
    worksheet.append(row)

worksheet = workbook.create_sheet("Active Alliance Avg Finish")
worksheet.append(
    ["Alliance", "Level", "Total Games", "Avg Finish", "Std Dev", "Pickrate"]
)
rows = [
    [
        name,
        level,
        len(active_alliance_finishes[(name, level)]),
        sum(active_alliance_finishes[(name, level)])
        / len(active_alliance_finishes[(name, level)]),
        stdev(active_alliance_finishes[(name, level)])
        if len(active_alliance_finishes[(name, level)]) > 1
        else None,
        len(active_alliance_finishes[(name, level)]) / total_games,
    ]
    for name, level in active_alliance_finishes
]

for row in sorted(rows, key=lambda x: x[2]):
    worksheet.append(row)

worksheet = workbook.create_sheet("Item Avg Finish")
worksheet.append(["Item", "Total Games", "Avg Finish", "Std Dev", "Pickrate"])
rows = [
    [
        item,
        len(item_finishes[item]),
        sum(item_finishes[item]) / len(item_finishes[item]),
        stdev(item_finishes[item]) if len(item_finishes[item]) > 1 else None,
        len(item_finishes[item]) / total_games,
    ]
    for item in item_finishes
]
print(rows)
for row in sorted(rows, key=lambda x: x[2]):
    worksheet.append(row)

worksheet = workbook.create_sheet("Hero Avg Finish by Item")
worksheet.append(
    ["Hero", "No Item"] + [item for item in sorted(item_finishes)]
)
rows = []
for hero in sorted(hero_finishes):
    rows.append(
        [hero, mean(hero_finishes[hero])]
        + [
            (lambda x: mean(x) if x else None)(
                hero_finishes_by_item.get(hero, {}).get(item, [])
            )
            for item in sorted(item_finishes)
        ]
    )

for row in rows:
    worksheet.append(row)


workbook.save(results_dir / "winrates.xlsx")
